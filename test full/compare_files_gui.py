"""
File Compare Studio v2 — flexible multi-file workspace
----------------------------------------------------------
Developed & Copyrighted by Tarek (Unlimited)
All Rights Reserved © 2026.

- Load as many files as you want (CSV / XLSX).
- Configure comparison key columns and dynamic keep-filters.
- Choose exactly which unique or common files to export.
- Fully asynchronous and lightweight background execution.

The original source files are only READ — never modified.

Install requirements:
    pip install customtkinter pandas openpyxl

Run:
    python compare_files_gui.py
"""

import os
import sys
import traceback
import subprocess
import threading
from tkinter import filedialog, messagebox

import customtkinter as ctk
import pandas as pd

# ----------------------------------------------------------------------------
# Theme / palette  (tuples are (light_mode, dark_mode))
# ----------------------------------------------------------------------------
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

BG            = ("#F2F3F9", "#0B0D14")
CARD_BG       = ("#FFFFFF", "#151823")
CARD_BORDER   = ("#E4E7F2", "#242838")
ROW_BG        = ("#F6F7FB", "#1C2030")
TEXT_PRIMARY  = ("#13151D", "#F5F6FA")
TEXT_SECOND   = ("#5B6175", "#9098B1")
ACCENT        = "#7C5CFF"
ACCENT_HOVER  = "#6A47F2"
TEAL          = "#16D6B5"
TEAL_HOVER    = "#10B89A"
DANGER        = "#FF5C7A"
DANGER_HOVER  = "#E84766"

FONT_FAMILY = "Segoe UI"


def F(size, weight="normal"):
    return ctk.CTkFont(family=FONT_FAMILY, size=size, weight=weight)


# ----------------------------------------------------------------------------
# Data helpers (pure logic, independent from the UI)
# ----------------------------------------------------------------------------
def read_file(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        last_err = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
            try:
                return pd.read_csv(path, dtype=str, keep_default_na=False, encoding=enc)
            except UnicodeDecodeError as e:
                last_err = e
                continue
        raise ValueError(f"Could not read '{path}' with common encodings.") from last_err
    elif ext in (".xlsx", ".xls"):
        return pd.read_excel(path, dtype=str)
    else:
        raise ValueError(f"Unsupported file extension '{ext}'. Use .csv, .xlsx or .xls")


def get_file_metadata(path: str):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        last_err = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
            try:
                df_header = pd.read_csv(path, nrows=0, encoding=enc)
                columns = list(df_header.columns)
                with open(path, 'rb') as f:
                    row_count = sum(1 for _ in f) - 1
                return columns, max(0, row_count)
            except Exception as e:
                last_err = e
                continue
        raise ValueError(f"Could not read CSV header: {last_err}")
    elif ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, keep_links=False)
        sheet = wb.active
        columns = []
        for row in sheet.iter_rows(max_row=1, values_only=True):
            columns = [str(c) for c in row if c is not None]
            break
        row_count = sheet.max_row
        if row_count is None:
            row_count = sum(1 for _ in sheet.iter_rows(values_only=True)) - 1
        else:
            row_count -= 1
        wb.close()
        return columns, max(0, row_count)
    else:
        raise ValueError(f"Unsupported file extension '{ext}'. Use .csv, .xlsx or .xls")


def guess_ip_column(columns):
    for c in columns:
        if "ip" in c.lower():
            return c
    return columns[0] if columns else ""


def run_comparison(df1, df2, key1, key2):
    """Returns (only_in_1, only_in_2, (matched_1, matched_2_with_match_key))."""
    d1, d2 = df1.copy(), df2.copy()
    d1["_match_key"] = d1[key1].astype(str).str.strip().str.lower()
    d2["_match_key"] = d2[key2].astype(str).str.strip().str.lower()

    set1, set2 = set(d1["_match_key"]), set(d2["_match_key"])
    only1, only2, common_keys = set1 - set2, set2 - set1, set1 & set2

    only_in_1 = d1[d1["_match_key"].isin(only1)].drop(columns=["_match_key"]).reset_index(drop=True)
    only_in_2 = d2[d2["_match_key"].isin(only2)].drop(columns=["_match_key"]).reset_index(drop=True)

    m1 = d1[d1["_match_key"].isin(common_keys)].reset_index(drop=True)
    m2 = d2[d2["_match_key"].isin(common_keys)].reset_index(drop=True)

    return only_in_1, only_in_2, (m1, m2)


# ----------------------------------------------------------------------------
# Small reusable UI bits
# ----------------------------------------------------------------------------
class StepHeader(ctk.CTkFrame):
    def __init__(self, master, number, title, subtitle=""):
        super().__init__(master, fg_color="transparent")
        bubble = ctk.CTkLabel(
            self, text=str(number), width=28, height=28, corner_radius=14,
            fg_color=ACCENT, text_color="#FFFFFF", font=F(13, "bold"),
        )
        bubble.grid(row=0, column=0, rowspan=2 if subtitle else 1, padx=(0, 10), sticky="n")
        ctk.CTkLabel(self, text=title, font=F(16, "bold"), text_color=TEXT_PRIMARY, anchor="w").grid(
            row=0, column=1, sticky="w"
        )
        if subtitle:
            ctk.CTkLabel(self, text=subtitle, font=F(12), text_color=TEXT_SECOND, anchor="w").grid(
                row=1, column=1, sticky="w"
            )


class Card(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(
            master, fg_color=CARD_BG, corner_radius=18, border_width=1,
            border_color=CARD_BORDER, **kwargs
        )


# ----------------------------------------------------------------------------
# One row in the file pool (step 1)
# ----------------------------------------------------------------------------
class FileRow(ctk.CTkFrame):
    def __init__(self, master, app, on_remove, on_changed):
        super().__init__(master, fg_color=ROW_BG, corner_radius=12)
        self.app = app
        self.on_remove = on_remove
        self.on_changed = on_changed
        self.path = None
        self.columns = []
        self.row_count = 0
        self.df = None
        
        self.grid_columnconfigure(2, weight=1)

        self.include_var = ctk.BooleanVar(value=True)
        self.include_cb = ctk.CTkCheckBox(
            self, text="", variable=self.include_var, width=20,
            command=self._on_include_toggled, fg_color=ACCENT
        )
        self.include_cb.grid(row=0, column=0, padx=(12, 4), pady=10, sticky="w")

        browse = ctk.CTkButton(
            self, text="Browse", width=70, corner_radius=10, fg_color=ACCENT,
            hover_color=ACCENT_HOVER, font=F(12, "bold"), command=self._browse,
        )
        browse.grid(row=0, column=1, padx=4, pady=10)

        self.entry = ctk.CTkEntry(
            self, placeholder_text="No file selected…", fg_color=("#FFFFFF", "#11141E"),
            border_color=CARD_BORDER, text_color=TEXT_PRIMARY, font=F(12),
        )
        self.entry.grid(row=0, column=2, padx=6, pady=10, sticky="ew")

        self.status = ctk.CTkLabel(self, text="—", font=F(11), text_color=TEXT_SECOND, width=150, anchor="w")
        self.status.grid(row=0, column=3, padx=6, pady=10)

        self.key_menu = ctk.CTkOptionMenu(
            self, values=["—"], width=130, fg_color=("#E9EBF6", "#262B3D"),
            button_color=ACCENT, button_hover_color=ACCENT_HOVER,
            text_color=TEXT_PRIMARY, font=F(12), command=self._on_key_selected
        )
        self.key_menu.grid(row=0, column=4, padx=6, pady=10)

        remove_btn = ctk.CTkButton(
            self, text="✕", width=32, height=32, corner_radius=10, fg_color="transparent",
            hover_color=DANGER, text_color=DANGER, font=F(14, "bold"),
            command=lambda: self.on_remove(self),
        )
        remove_btn.grid(row=0, column=5, padx=(4, 10), pady=10)

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select a file",
            filetypes=[("CSV / Excel files", "*.csv *.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path:
            return
        self.entry.delete(0, "end")
        self.entry.insert(0, path)
        self.status.configure(text="⏳ Loading...", text_color=ACCENT)
        self.key_menu.configure(values=["—"])
        self.key_menu.set("—")
        threading.Thread(target=self._load, args=(path,), daemon=True).start()

    def _load(self, path):
        try:
            cols, row_count = get_file_metadata(path)
            self.app.after(0, lambda: self._on_load_success(path, cols, row_count))
        except Exception as e:
            self.app.after(0, lambda: self._on_load_failure(path, str(e)))

    def _on_load_success(self, path, cols, row_count):
        self.path = path
        self.columns = cols
        self.row_count = row_count
        self.df = None
        self.status.configure(text=f"✓ {row_count} rows · {len(cols)} cols", text_color=TEAL)
        self.key_menu.configure(values=cols)
        guessed = guess_ip_column(cols)
        self.key_menu.set(guessed if guessed else cols[0])
        self.on_changed()

    def _on_load_failure(self, path, err_msg):
        self.columns = []
        self.row_count = 0
        self.df = None
        self.path = path
        self.status.configure(text=f"✕ Error", text_color=DANGER)
        self.key_menu.configure(values=["—"])
        self.key_menu.set("—")
        self.app._log(f"Error loading {os.path.basename(path)}: {err_msg}")
        self.on_changed()

    def _on_include_toggled(self):
        self.on_changed()

    def _on_key_selected(self, val):
        self.on_changed()

    def label(self, slot_number=None):
        if not self.path:
            return "No file loaded"
        return os.path.basename(self.path)
        return os.path.basename(self.path)


# ----------------------------------------------------------------------------
# One dynamic filter line: [file: A or B] [column] [values to keep] [remove]
# ----------------------------------------------------------------------------
class FilterRow(ctk.CTkFrame):
    def __init__(self, master, app, on_remove):
        super().__init__(master, fg_color=ROW_BG, corner_radius=12)
        self.app = app
        self.on_remove = on_remove
        self.grid_columnconfigure((0, 1, 2), weight=1)

        self.file_menu = ctk.CTkOptionMenu(
            self, values=app.included_labels() or ["—"], command=self._on_file_change,
            width=170, fg_color=("#E9EBF6", "#262B3D"), button_color=ACCENT,
            button_hover_color=ACCENT_HOVER, text_color=TEXT_PRIMARY, font=F(12),
        )
        self.file_menu.grid(row=0, column=0, padx=(12, 6), pady=12, sticky="ew")

        self.col_menu = ctk.CTkOptionMenu(
            self, values=["—"], width=180, fg_color=("#E9EBF6", "#262B3D"),
            button_color=ACCENT, button_hover_color=ACCENT_HOVER,
            text_color=TEXT_PRIMARY, font=F(12),
        )
        self.col_menu.grid(row=0, column=1, padx=6, pady=12, sticky="ew")

        self.values_entry = ctk.CTkEntry(
            self, placeholder_text="values to KEEP, comma-separated  e.g. BNG, Router, Switch, Controller",
            fg_color=("#FFFFFF", "#11141E"), border_color=CARD_BORDER, text_color=TEXT_PRIMARY, font=F(12),
        )
        self.values_entry.grid(row=0, column=2, padx=6, pady=12, sticky="ew")

        remove_btn = ctk.CTkButton(
            self, text="✕", width=32, height=32, corner_radius=10,
            fg_color="transparent", hover_color=DANGER, text_color=DANGER,
            font=F(14, "bold"), command=lambda: self.on_remove(self),
        )
        remove_btn.grid(row=0, column=3, padx=(6, 12), pady=12)

        self.refresh_file_options()

    def refresh_file_options(self):
        labels = self.app.included_labels()
        current = self.file_menu.get()
        self.file_menu.configure(values=labels or ["—"])
        if labels and current in labels:
            self.file_menu.set(current)
            self._on_file_change(current)
        elif labels:
            self.file_menu.set(labels[0])
            self._on_file_change(labels[0])
        else:
            self.file_menu.set("—")
            self.col_menu.configure(values=["—"])
            self.col_menu.set("—")

    def _on_file_change(self, label):
        cols = self.app.columns_for_label(label)
        self.col_menu.configure(values=cols if cols else ["—"])
        self.col_menu.set(cols[0] if cols else "—")

    def get_spec(self):
        return {
            "which_label": self.file_menu.get(),
            "column": self.col_menu.get(),
            "values": [v.strip() for v in self.values_entry.get().split(",") if v.strip()],
        }


# ----------------------------------------------------------------------------
# Main application
# ----------------------------------------------------------------------------
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("File Compare Studio")
        self.geometry("1120x920")
        self.minsize(940, 700)
        self.configure(fg_color=BG)

        self.file_rows = []
        self.filter_rows = []
        self.legacy_switches = {}  # path -> CTkSwitch
        self.output_dir = None

        self._build_header()
        self._build_footer()
        self._build_scroll_area()
        self._build_file_card()
        self._build_filter_card()
        self._build_output_card()
        self._build_log_card()

        # start with three empty file slots
        self._add_file_row()
        self._add_file_row()
        self._add_file_row()

    # ---------------- header ----------------
    def _build_header(self):
        bar = ctk.CTkFrame(self, fg_color="transparent", height=70)
        bar.pack(fill="x", padx=28, pady=(22, 6))

        left = ctk.CTkFrame(bar, fg_color="transparent")
        left.pack(side="left")
        ctk.CTkLabel(left, text="File Compare Studio", font=F(26, "bold"), text_color=TEXT_PRIMARY).pack(anchor="w")
        ctk.CTkLabel(
            left, text="Load files → configure keys → optional filters → export options → run",
            font=F(13), text_color=TEXT_SECOND,
        ).pack(anchor="w")

        right = ctk.CTkFrame(bar, fg_color="transparent")
        right.pack(side="right")
        self.mode_switch = ctk.CTkSegmentedButton(
            right, values=["Dark", "Light"], command=self._toggle_mode,
            selected_color=ACCENT, selected_hover_color=ACCENT_HOVER, font=F(12, "bold"),
        )
        self.mode_switch.set("Dark")
        self.mode_switch.pack(anchor="e")

    def _toggle_mode(self, value):
        ctk.set_appearance_mode("dark" if value == "Dark" else "light")

    def _build_footer(self):
        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.pack(fill="x", side="bottom", padx=28, pady=(0, 10))
        ctk.CTkLabel(
            footer, 
            text="© 2026 Tarek (Unlimited). All rights reserved. Unauthorized duplication or distribution is prohibited.", 
            font=F(11), text_color=TEXT_SECOND, anchor="center"
        ).pack(fill="x")

    # ---------------- scroll container ----------------
    def _build_scroll_area(self):
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=28, pady=(6, 22))
        self.scroll.grid_columnconfigure(0, weight=1)

    # ---------------- card 1: file pool ----------------
    def _build_file_card(self):
        card = Card(self.scroll)
        card.grid(row=0, column=0, sticky="ew", pady=(0, 18))
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=22, pady=20)
        inner.grid_columnconfigure(0, weight=1)

        StepHeader(
            inner, 1, "Load files & choose comparison keys",
            "Include files in comparison and select which column identifies matches (e.g. IP Address)",
        ).grid(row=0, column=0, sticky="w", pady=(0, 14))

        self.file_rows_frame = ctk.CTkFrame(inner, fg_color="transparent")
        self.file_rows_frame.grid(row=1, column=0, sticky="ew")
        self.file_rows_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkButton(
            inner, text="+ Add File", width=120, height=32, corner_radius=10,
            fg_color="transparent", border_width=1, border_color=ACCENT, text_color=ACCENT,
            hover_color=("#EFEBFF", "#1E1A33"), font=F(12, "bold"), command=self._add_file_row,
        ).grid(row=2, column=0, sticky="w", pady=(12, 0))

    def _add_file_row(self):
        row = FileRow(self.file_rows_frame, self, self._remove_file_row, self._on_files_changed)
        row.pack(fill="x", pady=6)
        self.file_rows.append(row)

    def _remove_file_row(self, row):
        row.destroy()
        self.file_rows.remove(row)
        self._on_files_changed()

    def _on_files_changed(self):
        for row in self.filter_rows:
            row.refresh_file_options()
        self._refresh_legacy_switches()
        self._refresh_export_choices()

    # ---------------- loaded-file lookup helpers ----------------
    def loaded_rows(self):
        return [r for r in self.file_rows if r.path is not None]

    def included_rows(self):
        return [r for r in self.file_rows if r.path is not None and r.include_var.get()]

    def included_labels(self):
        return [r.label() for r in self.included_rows()]

    def _row_for_label(self, label):
        for r in self.file_rows:
            if r.label() == label:
                return r
        return None

    def columns_for_label(self, label):
        row = self._row_for_label(label)
        return row.columns if row is not None else []

    # ---------------- card 2: filters ----------------
    def _build_filter_card(self):
        self.filter_card = Card(self.scroll)
        self.filter_card.grid(row=1, column=0, sticky="ew", pady=(0, 18))
        self.filter_inner = ctk.CTkFrame(self.filter_card, fg_color="transparent")
        self.filter_inner.pack(fill="x", padx=22, pady=20)
        self.filter_inner.grid_columnconfigure(0, weight=1)

        StepHeader(
            self.filter_inner, 2, "Filters (optional)",
            "Keep only rows whose column value matches what you list — applied before comparing",
        ).grid(row=0, column=0, sticky="w", pady=(0, 14))

        self.filter_rows_frame = ctk.CTkFrame(self.filter_inner, fg_color="transparent")
        self.filter_rows_frame.grid(row=1, column=0, sticky="ew")
        self.filter_rows_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkButton(
            self.filter_inner, text="+ Add Filter", width=130, height=32, corner_radius=10,
            fg_color="transparent", border_width=1, border_color=ACCENT, text_color=ACCENT,
            hover_color=("#EFEBFF", "#1E1A33"), font=F(12, "bold"), command=self._add_filter_row,
        ).grid(row=2, column=0, sticky="w", pady=(12, 4))

        self.legacy_frame = ctk.CTkFrame(self.filter_inner, fg_color="transparent")
        self.legacy_frame.grid(row=3, column=0, sticky="ew", pady=(14, 0))

    def _add_filter_row(self):
        row = FilterRow(self.filter_rows_frame, self, self._remove_filter_row)
        row.pack(fill="x", pady=6)
        self.filter_rows.append(row)

    def _remove_filter_row(self, row):
        row.destroy()
        self.filter_rows.remove(row)

    def _refresh_legacy_switches(self):
        for w in self.legacy_frame.winfo_children():
            w.destroy()
        self.legacy_switches = {}
        for r in self.included_rows():
            if r.columns and "ActivationDate" in r.columns:
                name = os.path.basename(r.path)
                sw = ctk.CTkSwitch(
                    self.legacy_frame,
                    text=f"Remove '1970' ActivationDate rows from {name}",
                    progress_color=ACCENT, font=F(12), text_color=TEXT_PRIMARY,
                )
                sw.pack(anchor="w", pady=4)
                self.legacy_switches[r.path] = sw

    # ---------------- card 3: output options + destination + run ----------------
    def _build_output_card(self):
        card = Card(self.scroll)
        card.grid(row=2, column=0, sticky="ew", pady=(0, 18))
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=22, pady=20)
        inner.grid_columnconfigure(0, weight=1)

        StepHeader(
            inner, 3, "Choose what to export",
            "Pick any combination of outputs to generate",
        ).grid(row=0, column=0, sticky="w", pady=(0, 14))

        self.export_choices_frame = ctk.CTkFrame(inner, fg_color="transparent")
        self.export_choices_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))

        self.out_common = ctk.CTkSwitch(inner, text="Export common rows (matched across all files)", progress_color=ACCENT,
                                         font=F(12), text_color=TEXT_PRIMARY)
        self.out_common.select()
        self.out_common.grid(row=2, column=0, sticky="w", pady=4)

        dest_row = ctk.CTkFrame(inner, fg_color="transparent")
        dest_row.grid(row=3, column=0, sticky="ew", pady=(16, 0))
        dest_row.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(dest_row, text="Save to", font=F(12, "bold"), text_color=TEXT_SECOND, width=60).grid(
            row=0, column=0, sticky="w"
        )
        self.dest_entry = ctk.CTkEntry(
            dest_row, placeholder_text="Choose a destination folder…", fg_color=("#FFFFFF", "#11141E"),
            border_color=CARD_BORDER, text_color=TEXT_PRIMARY, font=F(12),
        )
        self.dest_entry.grid(row=0, column=1, sticky="ew", padx=10)
        ctk.CTkButton(
            dest_row, text="Browse", width=90, corner_radius=10, fg_color=ACCENT,
            hover_color=ACCENT_HOVER, font=F(12, "bold"), command=self._browse_dest,
        ).grid(row=0, column=2)

        self.run_btn = ctk.CTkButton(
            inner, text="▶  Run Comparison", height=42, corner_radius=12, fg_color=ACCENT,
            hover_color=ACCENT_HOVER, font=F(14, "bold"), command=self._run,
        )
        self.run_btn.grid(row=4, column=0, sticky="ew", pady=(18, 0))

        self.progress_bar = ctk.CTkProgressBar(inner, progress_color=TEAL)
        self.progress_bar.grid(row=5, column=0, sticky="ew", pady=(10, 0))
        self.progress_bar.set(0)
        self.progress_bar.grid_remove()

    def _refresh_export_choices(self):
        for w in self.export_choices_frame.winfo_children():
            w.destroy()
        self.unique_export_switches = {}
        for r in self.included_rows():
            name = os.path.basename(r.path) if r.path else "No file loaded"
            sw = ctk.CTkSwitch(
                self.export_choices_frame,
                text=f"Export unique rows for {name} (only in this file)",
                progress_color=ACCENT, font=F(12), text_color=TEXT_PRIMARY,
            )
            sw.select()
            sw.pack(anchor="w", pady=4)
            self.unique_export_switches[r] = sw

    def _browse_dest(self):
        folder = filedialog.askdirectory(title="Choose destination folder")
        if folder:
            self.dest_entry.delete(0, "end")
            self.dest_entry.insert(0, folder)

    # ---------------- card 4: log / results ----------------
    def _build_log_card(self):
        card = Card(self.scroll)
        card.grid(row=3, column=0, sticky="ew")
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=22, pady=20)
        inner.grid_columnconfigure(0, weight=1)

        StepHeader(inner, 4, "Results").grid(row=0, column=0, sticky="w", pady=(0, 10))

        self.log_box = ctk.CTkTextbox(
            inner, height=160, fg_color=("#FFFFFF", "#11141E"), text_color=TEXT_PRIMARY,
            font=ctk.CTkFont(family="Consolas", size=12), corner_radius=10,
        )
        self.log_box.grid(row=1, column=0, sticky="ew")
        self.log_box.configure(state="disabled")

        self.open_folder_btn = ctk.CTkButton(
            inner, text="Open Output Folder", width=170, height=34, corner_radius=10,
            fg_color="transparent", border_width=1, border_color=TEAL, text_color=TEAL,
            hover_color=("#E6FBF6", "#102420"), font=F(12, "bold"), command=self._open_output_folder,
        )
        self.open_folder_btn.grid(row=2, column=0, sticky="w", pady=(12, 0))
        self.open_folder_btn.configure(state="disabled")

    def _log(self, text):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", text + "\n")
        self.log_box.configure(state="disabled")
        self.log_box.see("end")

    def _open_output_folder(self):
        if not self.output_dir:
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(self.output_dir)
            elif sys.platform == "darwin":
                subprocess.run(["open", self.output_dir])
            else:
                subprocess.run(["xdg-open", self.output_dir])
        except Exception as e:
            messagebox.showinfo("Output folder", f"Files were saved to:\n{self.output_dir}\n\n({e})")

    # ---------------- run logic ----------------
    def _run(self):
        included = self.included_rows()
        if len(included) < 2:
            messagebox.showwarning("Incomplete setup", "Please load and include at least two files to compare.")
            return

        unique_paths = [r.path for r, sw in self.unique_export_switches.items() if sw.get() == 1 and r.path]
        want_common = self.out_common.get() == 1
        if not (unique_paths or want_common):
            messagebox.showwarning("Nothing selected", "Pick at least one output to generate.")
            return

        dest = self.dest_entry.get().strip()
        if not dest:
            dest = os.path.dirname(os.path.abspath(included[0].path))
        if not os.path.isdir(dest):
            try:
                os.makedirs(dest, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Invalid destination", f"Couldn't use that folder:\n{e}")
                return
        self.output_dir = dest

        self.run_btn.configure(state="disabled", text="Comparing...")
        self.progress_bar.grid()
        self.progress_bar.configure(mode="indeterminate")
        self.progress_bar.start()

        threading.Thread(target=self._run_async, args=(included, unique_paths, want_common, dest), daemon=True).start()

    def _run_async(self, included, unique_paths, want_common, dest):
        try:
            # Load files on-demand in the background thread
            dfs = []
            for r in included:
                basename = os.path.splitext(os.path.basename(r.path))[0]
                self.after(0, lambda name=basename: self._log(f"Reading file '{name}'…"))
                dfs.append(read_file(r.path))

            paths = [r.path for r in included]
            basenames = [os.path.splitext(os.path.basename(p))[0] for p in paths]
            keys = [r.key_menu.get() for r in included]

            # 1. Apply generic keep-filters
            for row in self.filter_rows:
                spec = row.get_spec()
                for idx, r in enumerate(included):
                    if r.label() == spec["which_label"]:
                        col = spec["column"]
                        vals = spec["values"]
                        if not vals or col not in dfs[idx].columns:
                            continue
                        before = len(dfs[idx])
                        vals_lower = [v.lower() for v in vals]
                        mask = dfs[idx][col].astype(str).str.strip().str.lower().isin(vals_lower)
                        dfs[idx] = dfs[idx][mask].reset_index(drop=True)
                        self.after(0, lambda idx=idx, col=col, vals=vals, before=before, after_len=len(dfs[idx]): 
                            self._log(f"Filter: File {basenames[idx]} · {col} kept {vals} → {before} → {after_len} rows"))

            # 2. Apply 1970 ActivationDate filters
            for idx, r in enumerate(included):
                if r.path in self.legacy_switches and self.legacy_switches[r.path].get() == 1:
                    df = dfs[idx]
                    if "ActivationDate" in df.columns:
                        before = len(df)
                        mask = ~df["ActivationDate"].astype(str).str.strip().str.startswith("1970")
                        dfs[idx] = df[mask].reset_index(drop=True)
                        self.after(0, lambda idx=idx, before=before, after_len=len(dfs[idx]):
                            self._log(f"Filter: removed 1970 ActivationDate rows in {basenames[idx]} → {before} → {after_len} rows"))

            # 3. Add normalized match keys and build key sets
            for idx, r in enumerate(included):
                key = keys[idx]
                if key not in dfs[idx].columns:
                    raise ValueError(f"Key column '{key}' not found in {basenames[idx]}")
                dfs[idx]["_match_key"] = dfs[idx][key].astype(str).str.strip().str.lower()

            sets = [set(df["_match_key"]) for df in dfs]

            self.after(0, lambda: self._log("—" * 50))

            # 4. Generate unique outputs
            if unique_paths:
                for idx, r in enumerate(included):
                    if r.path not in unique_paths:
                        continue
                    df = dfs[idx]
                    set_i = sets[idx]
                    other_sets = [sets[j] for j in range(len(sets)) if j != idx]
                    union_others = set.union(*other_sets) if other_sets else set()
                    unique_keys = set_i - union_others
                    only_in_df = df[df["_match_key"].isin(unique_keys)].drop(columns=["_match_key"]).reset_index(drop=True)
                    
                    out_path = os.path.join(dest, f"only_in_{basenames[idx]}.xlsx")
                    only_in_df.to_excel(out_path, index=False)
                    self.after(0, lambda name=basenames[idx], count=len(only_in_df), path=out_path:
                        self._log(f"Only in {name}: {count} rows → {path}"))

            # 5. Generate common outputs
            if want_common:
                common_keys = set.intersection(*sets) if sets else set()
                common_df = None
                for idx, r in enumerate(included):
                    matched = dfs[idx][dfs[idx]["_match_key"].isin(common_keys)].copy()
                    rename_dict = {col: f"{col}_{basenames[idx]}" for col in matched.columns if col != "_match_key"}
                    matched = matched.rename(columns=rename_dict)
                    if common_df is None:
                        common_df = matched
                    else:
                        common_df = pd.merge(common_df, matched, on="_match_key")

                if common_df is not None and "_match_key" in common_df.columns:
                    common_df = common_df.drop(columns=["_match_key"]).reset_index(drop=True)
                else:
                    common_df = pd.DataFrame()

                out_path = os.path.join(dest, f"common_all_files.xlsx")
                common_df.to_excel(out_path, index=False)
                self.after(0, lambda count=len(common_df), path=out_path:
                    self._log(f"Matched in all: {count} rows → {path}"))

            self.after(0, lambda: self._log("Done. Original files were not modified."))
            self.after(0, self._run_success)

        except Exception as e:
            tb = traceback.format_exc()
            self.after(0, lambda: self._run_failed(tb))

    def _run_success(self):
        self.run_btn.configure(state="normal", text="▶  Run Comparison")
        self.progress_bar.stop()
        self.progress_bar.grid_remove()
        self.open_folder_btn.configure(state="normal")

    def _run_failed(self, tb):
        self.run_btn.configure(state="normal", text="▶  Run Comparison")
        self.progress_bar.stop()
        self.progress_bar.grid_remove()
        messagebox.showerror("Error while running comparison", tb)


class SplashScreen(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Loading...")
        # Borderless window
        self.overrideredirect(True)
        self.configure(fg_color="#0B0D14")
        
        # Center the window
        width, height = 550, 330
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Main content
        frame = ctk.CTkFrame(self, fg_color="transparent")
        frame.pack(expand=True, fill="both", padx=40, pady=40)
        
        ctk.CTkLabel(
            frame, text="File Compare Studio", 
            font=ctk.CTkFont(family="Segoe UI", size=32, weight="bold"), 
            text_color="#7C5CFF"
        ).pack(pady=(20, 5))
        
        ctk.CTkLabel(
            frame, text="Developed by Tarek (Unlimited)", 
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="normal"), 
            text_color="#9098B1"
        ).pack(pady=(0, 30))
        
        self.progress = ctk.CTkProgressBar(frame, progress_color="#16D6B5", width=400, height=8)
        self.progress.pack(pady=10)
        self.progress.set(0)
        
        self.status = ctk.CTkLabel(
            frame, text="Initializing workspace...", 
            font=ctk.CTkFont(family="Segoe UI", size=11), 
            text_color="#5B6175"
        )
        self.status.pack(pady=(10, 0))
        
        # Start progress loading loop
        self.progress_val = 0.0
        self._update_progress()
        
    def _update_progress(self):
        if self.progress_val < 1.0:
            self.progress_val += 0.04
            self.progress.set(self.progress_val)
            self.after(100, self._update_progress)
        else:
            self.destroy()
            app = App()
            app.mainloop()


if __name__ == "__main__":
    splash = SplashScreen()
    splash.mainloop()
